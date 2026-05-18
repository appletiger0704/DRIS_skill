#!/usr/bin/env python3
"""Extract food index from DRIs_template.xlsx.

Usage:
  python scripts/extract_food_index.py assets/DRIs_template.xlsx assets/food_database_index.json
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


FOOD_SHEET = "台灣食品成分表2020版"


def extract(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if FOOD_SHEET not in wb.sheetnames:
        raise ValueError(f"Workbook missing sheet: {FOOD_SHEET}")
    ws = wb[FOOD_SHEET]
    foods = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sample_name = row[0]
        if not sample_name:
            continue
        category = row[1] if len(row) > 1 else None
        item = {
            "sample_name": str(sample_name),
            "category": str(category) if category else "",
            "row": row_idx,
            "search_text": " ".join(str(v) for v in row[:8] if v not in (None, "")),
        }
        foods.append(item)
    return {"source": xlsx_path.name, "sheet": FOOD_SHEET, "count": len(foods), "foods": foods}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("output_json")
    args = parser.parse_args()
    data = extract(Path(args.xlsx_path))
    Path(args.output_json).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {data['count']} foods to {args.output_json}")


if __name__ == "__main__":
    main()
