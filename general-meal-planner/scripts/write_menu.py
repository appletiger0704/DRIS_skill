#!/usr/bin/env python3
"""Write a structured meal plan into the 計算 sheet of a DRIs Excel workbook.

Expected JSON schema:
{
  "menu_name": "example",
  "meals": [
    {
      "meal": "2午餐",
      "dishes": [
        {
          "dish_category": "主食",
          "dish_name": "白飯",
          "ingredients": [
            {"colloquial": "白米", "db_name": "粳米平均值", "weight_g": 50}
          ]
        }
      ]
    }
  ]
}

Lunch and dinner must contain exactly:
- 1 主食
- 1 主菜 or 主餐
- 3 副菜
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


VALID_MAIN_DISH = {"主菜", "主餐"}
MEAL_ORDER = {
    "1早餐": 10,
    "4早點": 20,
    "2午餐": 30,
    "5午點": 40,
    "3晚餐": 50,
    "6晚點": 60,
    "7流灌": 70,
}
DISH_ORDER = {
    "主食": 10,
    "主菜": 20,
    "主餐": 20,
    "副菜": 30,
    "調味料": 90,
    "補充食材": 95,
}


def canonical_meal(meal_name: str) -> str:
    text = str(meal_name or "")
    for key in MEAL_ORDER:
        if text.startswith(key):
            return key
    return text


def meal_sort_key(meal_name: str) -> int:
    return MEAL_ORDER.get(canonical_meal(meal_name), 999)


def dish_sort_key(dish: dict[str, Any], index: int) -> tuple[int, int, str]:
    category = str(dish.get("dish_category", ""))
    name = str(dish.get("dish_name", ""))
    order = DISH_ORDER.get(category, 50)
    # Keep 副菜1/副菜2/副菜3 order if embedded in dish name.
    if "副菜1" in name:
        order = 31
    elif "副菜2" in name:
        order = 32
    elif "副菜3" in name:
        order = 33
    return (order, index, name)


def validate_lunch_dinner_structure(menu: dict[str, Any]) -> None:
    errors = []
    for meal in menu.get("meals", []):
        meal_name = canonical_meal(str(meal.get("meal", "")))
        if meal_name not in {"2午餐", "3晚餐"}:
            continue

        categories = [dish.get("dish_category") for dish in meal.get("dishes", [])]
        counts = Counter(categories)
        main_count = sum(counts[c] for c in VALID_MAIN_DISH)

        if counts.get("主食", 0) != 1 or main_count != 1 or counts.get("副菜", 0) != 3:
            errors.append(
                f"{meal_name} 結構錯誤：需要 1 主食 + 1 主菜/主餐 + 3 副菜；目前={dict(counts)}"
            )

    if errors:
        raise ValueError("\n".join(errors))


def iter_sorted_rows(menu: dict[str, Any]):
    meals = sorted(menu.get("meals", []), key=lambda m: meal_sort_key(str(m.get("meal", ""))))
    for meal in meals:
        meal_name = canonical_meal(str(meal.get("meal", "")))
        dishes = list(enumerate(meal.get("dishes", [])))
        dishes.sort(key=lambda pair: dish_sort_key(pair[1], pair[0]))
        for _, dish in dishes:
            dish_name = dish.get("dish_name")
            first = True
            for ing in dish.get("ingredients", []):
                yield {
                    "meal": meal_name,
                    "dish_name": dish_name if first else None,
                    "colloquial": ing.get("colloquial"),
                    "db_name": ing.get("db_name"),
                    "weight_g": float(ing.get("weight_g", 0)),
                }
                first = False


def validate_written_meal_order(ws) -> None:
    seen_orders = []
    for row in range(6, min(ws.max_row, 90) + 1):
        val = ws.cell(row=row, column=1).value
        if not val:
            continue
        seen_orders.append(meal_sort_key(str(val)))
    if seen_orders != sorted(seen_orders):
        raise ValueError("計算 sheet A欄餐別順序錯誤：需依早餐、早點、午餐、午點、晚餐、晚點、流灌排序")


def write_menu_to_sheet(xlsx_path: Path, menu_json: Path, output_path: Path) -> None:
    with menu_json.open("r", encoding="utf-8") as f:
        menu = json.load(f)

    validate_lunch_dinner_structure(menu)

    wb = load_workbook(xlsx_path)
    if "計算" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 計算 sheet")
    ws = wb["計算"]

    if menu.get("menu_name"):
        ws["D3"] = menu.get("menu_name")

    # Clear previous input rows conservatively.
    for row in range(6, min(ws.max_row, 90) + 1):
        for col in [1, 2, 3, 4, 8]:
            ws.cell(row=row, column=col).value = None

    row = 6
    for item in iter_sorted_rows(menu):
        if row > 90:
            raise ValueError("菜單食材超過 計算!A6:H90 可寫入範圍")
        ws.cell(row=row, column=1).value = item["meal"]
        ws.cell(row=row, column=2).value = item["dish_name"]
        ws.cell(row=row, column=3).value = item["colloquial"]
        ws.cell(row=row, column=4).value = item["db_name"]
        ws.cell(row=row, column=8).value = item["weight_g"]
        row += 1

    validate_written_meal_order(ws)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("menu_json")
    parser.add_argument("output_path")
    args = parser.parse_args()
    write_menu_to_sheet(Path(args.xlsx_path), Path(args.menu_json), Path(args.output_path))


if __name__ == "__main__":
    main()
