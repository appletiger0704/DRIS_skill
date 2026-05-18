#!/usr/bin/env python3
"""Check therapeutic menu nutrition, dish composition, and seasoning quality."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

NUTRIENT_COLS = {
    "calories_kcal": 8, "protein_g": 11, "fat_g": 12, "carb_g": 15, "fiber_g": 16,
    "sodium_mg": 24, "potassium_mg": 25, "calcium_mg": 26, "magnesium_mg": 27,
    "iron_mg": 28, "zinc_mg": 29, "phosphorus_mg": 30, "vitA_RE_ug": 34,
    "vitD_ug": 39, "vitE_TE_mg": 43, "vitB1_mg": 51, "vitB2_mg": 52,
    "niacin_mg": 53, "vitB6_mg": 54, "vitB12_ug": 55, "folate_ug": 56, "vitC_mg": 57,
}
DEFAULT_LOWER_CHECKS = ["calories_kcal", "calcium_mg", "iron_mg", "zinc_mg", "vitD_ug"]
DEFAULT_UPPER_CHECKS = ["protein_g", "sodium_mg", "potassium_mg", "phosphorus_mg"]
SEASONING_WORDS = {
    "鹽", "醬油", "醋", "米醋", "烏醋", "檸檬", "檸檬汁", "胡椒", "白胡椒", "黑胡椒",
    "五香", "咖哩", "肉桂", "八角", "辣椒", "蒜", "大蒜", "蒜粉", "薑", "老薑", "嫩薑", "薑粉",
    "薑黃", "青蔥", "蔥", "洋蔥", "九層塔", "羅勒", "薄荷", "洋香菜", "迷迭香", "香油",
    "芝麻油", "黑芝麻油", "芝麻", "黑芝麻", "冰糖", "砂糖", "糖", "蜂蜜",
}
NON_CORE_WORDS = {
    "油", "大豆油", "橄欖油", "香油", "芝麻油", "黑芝麻油", "調合芝麻油",
    "糖", "冰糖", "砂糖", "蜂蜜", "米醋", "醋", "檸檬汁", "白胡椒粉", "黑胡椒粉", "胡椒粉",
    "五香粉", "咖哩粉", "肉桂粉", "八角", "辣椒粉", "迷迭香粉", "羅勒片", "洋香菜片",
    "薑粉", "薑黃粉", "蒜粉", "鹽", "醬油", "低鈉鹽", "味精",
}
EXEMPT_CATEGORIES = {"調味料", "補充食材"}


def num(value: Any) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def load_foods(template_xlsx: Path) -> dict[str, dict[str, float]]:
    wb = load_workbook(template_xlsx, data_only=True, read_only=True)
    if "台灣食品成分表2020版" not in wb.sheetnames:
        raise ValueError("template must contain 台灣食品成分表2020版")
    ws = wb["台灣食品成分表2020版"]
    max_col = max(NUTRIENT_COLS.values())
    foods: dict[str, dict[str, float]] = {}
    for row in ws.iter_rows(min_row=3, max_col=max_col, values_only=True):
        name = row[0]
        if not name:
            continue
        foods[str(name)] = {key: num(row[col - 1]) for key, col in NUTRIENT_COLS.items()}
    return foods


def calc_menu(menu: dict[str, Any], foods: dict[str, dict[str, float]]) -> dict[str, float]:
    totals = {key: 0.0 for key in NUTRIENT_COLS}
    missing: list[str] = []
    for meal in menu.get("meals", []):
        for dish in meal.get("dishes", []):
            for ing in dish.get("ingredients", []):
                db_name = str(ing.get("db_name", ""))
                weight = num(ing.get("weight_g"))
                if db_name not in foods:
                    missing.append(db_name)
                    continue
                for key, value in foods[db_name].items():
                    totals[key] += value * weight / 100.0
    if missing:
        raise ValueError("food names not found in official database: " + ", ".join(sorted(set(missing))))
    return totals


def load_targets(target_json: Path) -> dict[str, float]:
    payload = json.loads(target_json.read_text(encoding="utf-8"))
    raw = payload.get("targets", payload)
    targets = {key: num(value) for key, value in raw.items() if value is not None}
    if "vitD_IU" in targets and "vitD_ug" not in targets:
        targets["vitD_ug"] = targets["vitD_IU"] / 40.0
    if "vitD_ug" in targets and "vitD_IU" not in targets:
        targets["vitD_IU"] = targets["vitD_ug"] * 40.0
    return targets


def check_quality(totals: dict[str, float], targets: dict[str, float], strict: bool) -> list[str]:
    messages: list[str] = []
    lower_tol = 0.95 if strict else 0.90
    kcal_tol = 0.03
    protein_upper_factor = 1.10 if strict else 1.20
    for key in DEFAULT_LOWER_CHECKS:
        if key not in targets or targets[key] <= 0:
            continue
        if key == "calories_kcal":
            low = targets[key] * (1 - kcal_tol)
            high = targets[key] * (1 + kcal_tol)
            if not low <= totals[key] <= high:
                messages.append(f"{key} {totals[key]:.2f} outside ±{kcal_tol:.0%} of target {targets[key]:.2f}")
        elif totals[key] < targets[key] * lower_tol:
            messages.append(f"{key} {totals[key]:.2f} below {lower_tol:.0%} of target {targets[key]:.2f}")
    for key in DEFAULT_UPPER_CHECKS:
        if key not in targets or targets[key] <= 0:
            continue
        factor = protein_upper_factor if key == "protein_g" else 1.0
        if totals[key] > targets[key] * factor:
            messages.append(f"{key} {totals[key]:.2f} exceeds limit {targets[key] * factor:.2f}")
    return messages


def is_seasoning(ing: dict[str, Any]) -> bool:
    text = f"{ing.get('colloquial', '')} {ing.get('db_name', '')}"
    return any(word in text for word in SEASONING_WORDS)


def is_core_ingredient(ing: dict[str, Any]) -> bool:
    text = f"{ing.get('colloquial', '')} {ing.get('db_name', '')}"
    if any(word in text for word in NON_CORE_WORDS):
        return False
    return num(ing.get("weight_g")) > 0


def check_dish_composition(menu: dict[str, Any], *, min_core: int = 2, min_seasonings: int = 3) -> list[str]:
    messages: list[str] = []
    for meal in menu.get("meals", []):
        meal_name = meal.get("meal", "")
        for dish in meal.get("dishes", []):
            category = str(dish.get("dish_category", ""))
            if category in EXEMPT_CATEGORIES or dish.get("quality_exemption"):
                continue
            ingredients = dish.get("ingredients", [])
            core_count = sum(1 for ing in ingredients if is_core_ingredient(ing))
            seasoning_count = sum(1 for ing in ingredients if is_seasoning(ing))
            if core_count < min_core:
                messages.append(f"{meal_name} {dish.get('dish_name')} 主要食材不足：至少 {min_core} 種，目前 {core_count} 種")
            if seasoning_count < min_seasonings:
                messages.append(f"{meal_name} {dish.get('dish_name')} 調味/增味元素不足：至少 {min_seasonings} 種，目前 {seasoning_count} 種")
            if not dish.get("method"):
                messages.append(f"{meal_name} {dish.get('dish_name')} 缺少 method 作法簡述")
    return messages


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("template_xlsx")
    parser.add_argument("menu_json")
    parser.add_argument("targets_json")
    parser.add_argument("--strict", action="store_true", help="Use 95 percent lower threshold for Ca/Fe/Zn/VitD and tighter protein upper bound")
    parser.add_argument("--check-dish-composition", action="store_true", help="Require each dish to have >=2 core ingredients and >=3 seasonings unless exempted")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    menu = json.loads(Path(args.menu_json).read_text(encoding="utf-8"))
    foods = load_foods(Path(args.template_xlsx))
    targets = load_targets(Path(args.targets_json))
    totals = calc_menu(menu, foods)
    messages = check_quality(totals, targets, args.strict)
    if args.check_dish_composition:
        messages.extend(check_dish_composition(menu))
    result = {"totals": {k: round(v, 2) for k, v in totals.items()}, "messages": messages}
    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        for key in sorted(totals):
            print(f"{key}: {totals[key]:.2f}")
        if messages:
            print("Quality check failed:")
            for msg in messages:
                print(f"- {msg}")
        else:
            print("Quality check passed.")
    return 1 if messages else 0


if __name__ == "__main__":
    sys.exit(main())
