#!/usr/bin/env python3
"""Set custom DRIs target values in row 4 of the 與DRIs比較 sheet.

Usage:
  python set_dris_targets.py <xlsx_path> --targets targets.json
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


# Column mapping for the standard DRIs workbook.
TARGET_COLS = {
    "calories_kcal": "B",
    "protein_g": "E",
    "fat_g": "F",
    "carb_g": "I",
    "fiber_g": "J",
    "sodium_mg": "R",
    "potassium_mg": "S",
    "calcium_mg": "T",
    "iron_mg": "V",
    "phosphorus_mg": "X",
    "vitA_RE_ug": "AB",
    "vitD_IU": "AF",
    "vitD_ug": "AG",
    "vitE_TE_mg": "AK",
    "vitK1_ug": "AP",
    "vitK2_mk4_ug": "AQ",
    "vitK2_mk7_ug": "AR",
    "vitB1_mg": "AS",
    "vitB2_mg": "AT",
    "niacin_mg": "AU",
    "vitB6_mg": "AV",
    "vitB12_ug": "AW",
    "folate_ug": "AX",
    "vitC_mg": "AY",
}


def set_targets(xlsx_path: Path, targets_json: Path) -> None:
    with targets_json.open("r", encoding="utf-8") as f:
        payload = json.load(f)

    targets = payload.get("targets", payload)
    wb = load_workbook(xlsx_path)
    if "與DRIs比較" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 與DRIs比較 sheet")
    ws = wb["與DRIs比較"]

    # Convert vitamin D automatically if only one unit is provided.
    if "vitD_ug" in targets and "vitD_IU" not in targets:
        targets["vitD_IU"] = float(targets["vitD_ug"]) * 40
    if "vitD_IU" in targets and "vitD_ug" not in targets:
        targets["vitD_ug"] = float(targets["vitD_IU"]) / 40

    for key, col in TARGET_COLS.items():
        ws[f"{col}4"] = targets.get(key)

    if "patient_label" in payload:
        ws["A4"] = payload["patient_label"]

    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--targets", required=True)
    args = parser.parse_args()
    set_targets(Path(args.xlsx_path), Path(args.targets))


if __name__ == "__main__":
    main()
