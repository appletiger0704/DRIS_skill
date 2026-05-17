#!/usr/bin/env python3
"""Read computed nutrition totals from a recalculated DRIs Excel.

The workbook must be recalculated first, preferably with scripts/recalc.py.
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


TOTAL_ROW = 12
TARGET_ROW = 4
COLS = {
    "calories_kcal": "B",
    "protein_g": "E",
    "fat_g": "F",
    "carb_g": "I",
    "fiber_g": "J",
    "sodium_mg": "R",
    "potassium_mg": "S",
    "calcium_mg": "T",
    "iron_mg": "V",
    "phosphorus_mg": "X",
    "vitA_RE_ug": "AB",
    "vitD_IU": "AF",
    "vitD_ug": "AG",
    "vitE_TE_mg": "AK",
    "vitK1_ug": "AP",
    "vitK2_mk4_ug": "AQ",
    "vitK2_mk7_ug": "AR",
    "vitB1_mg": "AS",
    "vitB2_mg": "AT",
    "niacin_mg": "AU",
    "vitB6_mg": "AV",
    "vitB12_ug": "AW",
    "folate_ug": "AX",
    "vitC_mg": "AY",
}


def read_nutrition(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    if "與DRIs比較" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 與DRIs比較 sheet")
    ws = wb["與DRIs比較"]

    result = {"targets": {}, "totals": {}, "ratio": {}}
    for key, col in COLS.items():
        target = ws[f"{col}{TARGET_ROW}"].value
        total = ws[f"{col}{TOTAL_ROW}"].value
        result["targets"][key] = target
        result["totals"][key] = total
        try:
            result["ratio"][key] = None if not target else float(total) / float(target)
        except Exception:
            result["ratio"][key] = None

    # Vitamin K total from K1 + K2 MK-4 + K2 MK-7 if available.
    k_values = [result["totals"].get(k) or 0 for k in ["vitK1_ug", "vitK2_mk4_ug", "vitK2_mk7_ug"]]
    result["totals"]["vitK_total_ug"] = sum(k_values)
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    data = read_nutrition(Path(args.xlsx_path))
    if args.json:
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        for key, value in data["totals"].items():
            print(f"{key}: {value}")


if __name__ == "__main__":
    main()
